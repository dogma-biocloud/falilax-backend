import csv
import io
from typing import Any

from openpyxl import load_workbook


def parse_csv_bytes(file_bytes: bytes) -> list[dict[str, Any]]:
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def parse_xlsx_bytes(file_bytes: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    parsed_rows: list[dict[str, Any]] = []

    for row in rows[1:]:
        row_dict: dict[str, Any] = {}
        for idx, value in enumerate(row):
            header = headers[idx] if idx < len(headers) else f"column_{idx}"
            if header:
                row_dict[header] = value
        parsed_rows.append(row_dict)

    return parsed_rows